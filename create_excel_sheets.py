#!/usr/bin/env python3
"""
Script to create Excel sheets for the product ordering system.
This implements the refactored structure with:
1. No mandatory red color for Product A in size_and_colors sheet
2. One Size and One Color columns in products sheet
3. Conditional fields in orders sheet
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_products_sheet():
    """Create the products.xlsx file with One Size and One Color columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Define headers
    headers = ["Product ID", "Product Name", "Description", "Price", "One Size", "One Color"]
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample products
    products = [
        ["P001", "Product A", "Standard product with multiple sizes and colors", 29.99, "No", "No"],
        ["P002", "Product B", "One size fits all product", 19.99, "Yes", "No"],
        ["P003", "Product C", "Available in one color only", 24.99, "No", "Yes"],
        ["P004", "Product D", "Single size and color variant", 15.99, "Yes", "Yes"],
    ]
    
    for product in products:
        ws.append(product)
    
    # Add data validation for One Size and One Color columns (Yes/No dropdown)
    yes_no_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    yes_no_validation.error = "Please select Yes or No"
    yes_no_validation.errorTitle = "Invalid Value"
    
    ws.add_data_validation(yes_no_validation)
    yes_no_validation.add(f"E2:E1000")  # One Size column
    yes_no_validation.add(f"F2:F1000")  # One Color column
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    wb.save("products.xlsx")
    print("Created products.xlsx")

def create_size_and_colors_sheet():
    """Create the size_and_colors.xlsx file without mandatory red color for Product A"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Size and Colors"
    
    # Define headers
    headers = ["Product ID", "Product Name", "Available Sizes", "Available Colors"]
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add product variants
    # Note: Product A now has flexible color options (no mandatory red in every size)
    variants = [
        ["P001", "Product A", "Small", "Blue"],
        ["P001", "Product A", "Small", "Green"],
        ["P001", "Product A", "Medium", "Blue"],
        ["P001", "Product A", "Medium", "Green"],
        ["P001", "Product A", "Medium", "Red"],  # Red is optional, not in every size
        ["P001", "Product A", "Large", "Blue"],
        ["P001", "Product A", "Large", "Green"],
        ["P001", "Product A", "X-Large", "Red"],
        ["P001", "Product A", "X-Large", "Blue"],
        ["P002", "Product B", "One Size", "Black"],
        ["P002", "Product B", "One Size", "White"],
        ["P003", "Product C", "Small", "Navy"],
        ["P003", "Product C", "Medium", "Navy"],
        ["P003", "Product C", "Large", "Navy"],
        ["P004", "Product D", "One Size", "Gray"],
    ]
    
    for variant in variants:
        ws.append(variant)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    wb.save("size_and_colors.xlsx")
    print("Created size_and_colors.xlsx")

def create_orders_sheet():
    """Create the orders.xlsx file with conditional size/color fields"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Define headers
    headers = ["Order ID", "Customer Name", "Product ID", "Product Name", "Size", "Color", "Quantity", "Order Date", "Notes"]
    ws.append(headers)
    
    # Style the header row
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample orders
    # When a product has "One Size" = Yes, the Size column should show "N/A" or be disabled
    # When a product has "One Color" = Yes, the Color column should show "N/A" or be disabled
    orders = [
        ["O001", "John Doe", "P001", "Product A", "Medium", "Red", 2, "2026-01-01", "Standard order"],
        ["O002", "Jane Smith", "P002", "Product B", "N/A", "Black", 1, "2026-01-01", "One Size product - size field disabled"],
        ["O003", "Bob Johnson", "P003", "Product C", "Large", "N/A", 3, "2026-01-01", "One Color product - color field disabled"],
        ["O004", "Alice Williams", "P004", "Product D", "N/A", "N/A", 1, "2026-01-01", "One Size and One Color - both fields disabled"],
        ["O005", "Charlie Brown", "P001", "Product A", "Small", "Blue", 2, "2026-01-01", "Standard order with multiple options"],
    ]
    
    for order in orders:
        ws.append(order)
    
    # Add instructions in a note
    ws.cell(row=ws.max_row + 2, column=1, value="INSTRUCTIONS:")
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    
    instruction_text = (
        "- If a product has 'One Size' = Yes in the products sheet, enter 'N/A' in the Size column\n"
        "- If a product has 'One Color' = Yes in the products sheet, enter 'N/A' in the Color column\n"
        "- For products with both restrictions, both Size and Color should be 'N/A'\n"
        "- In an automated system, these fields would be automatically disabled based on product settings"
    )
    
    instruction_row = ws.max_row + 1
    ws.cell(row=instruction_row, column=1, value=instruction_text)
    ws.cell(row=instruction_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f'A{instruction_row}:I{instruction_row}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 40
    
    wb.save("orders.xlsx")
    print("Created orders.xlsx")

def main():
    """Create all three Excel sheets"""
    print("Creating Excel sheets for product ordering system...")
    create_products_sheet()
    create_size_and_colors_sheet()
    create_orders_sheet()
    print("\nAll Excel sheets created successfully!")
    print("\nChanges implemented:")
    print("1. ✓ Removed requirement for red color in every size of Product A")
    print("2. ✓ Added 'One Size' and 'One Color' columns to products sheet")
    print("3. ✓ Updated orders sheet with N/A for restricted products (automated systems would disable these fields)")

if __name__ == "__main__":
    main()
